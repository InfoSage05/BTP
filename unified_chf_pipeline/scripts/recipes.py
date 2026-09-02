"""
recipes.py
----------
One loader function per raw CHF source file. Each loader reads a single raw
file from data/raw/ and returns a DataFrame already translated into the
canonical column names/units defined in merge_datasets.CANONICAL_COLUMNS.

Design decisions (see project discussion for the full reasoning):
- Only point-level, real-experiment sources are loaded here. Range/summary
  tables (e.g. narrowchannel_chf_ml2025_source_summary.csv) and the smoothed
  Groeneveld LUT grid (data/chf_long_clean.csv) are intentionally excluded --
  they are not independent data points.
- CHF Dataset.csv is intentionally NOT loaded: it is a byte-identical
  duplicate of pinfin_chf_water_fc72.csv (verified during EDA).
- Any value that had to be assumed rather than read from the source file
  (e.g. pinfin's pressure, zhao2020's fluid) is recorded in the free-text
  `assumptions` column rather than silently hard-coded.
"""
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
DATA_RAW = ROOT / "data" / "raw"
DATA_EXTERNAL = DATA_RAW / "external"


def load_nrc_groeneveld() -> pd.DataFrame:
    df = pd.read_csv(DATA_EXTERNAL / "nrc_groeneveld_24579pt_chf_database.csv",
                      encoding="utf-8-sig", skiprows=[1])
    numeric_like = ["Number", "Tube Diameter", "Heated Length", "Pressure", "Mass Flux",
                     "Outlet Quality", "Inlet Subcooling", "Inlet Temperature", "CHF"]
    for c in numeric_like:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return pd.DataFrame({
        "source_row_id": df["Number"].astype(str),
        "geometry_family": "tube",
        "data_type": "raw_experimental",
        "fluid": "water",
        "pressure_kPa": df["Pressure"],
        "mass_flux_kg_m2s": df["Mass Flux"],
        "quality": df["Outlet Quality"],
        "subcooling_kJkg": df["Inlet Subcooling"],
        "diameter_mm": df["Tube Diameter"] * 1000,
        "heated_length_mm": df["Heated Length"] * 1000,
        "inlet_temperature_C": df["Inlet Temperature"],
        "reference_id": df["Reference ID"],
        "CHF_kW_m2": df["CHF"],
    })


def load_zhao2020() -> pd.DataFrame:
    df = pd.read_csv(DATA_EXTERNAL / "zhao2020_chf_flowboiling_tubes.csv", encoding="utf-8-sig")
    return pd.DataFrame({
        "source_row_id": df["id"].astype(str),
        "geometry_family": df["geometry"],
        "data_type": "raw_experimental",
        "fluid": "water",
        "author": df["author"],
        "pressure_kPa": df["pressure [MPa]"] * 1000,
        "mass_flux_kg_m2s": df["mass_flux [kg/m2-s]"],
        "quality": df["x_e_out [-]"],
        "diameter_mm": df["D_h [mm]"],
        "D_e_mm": df["D_e [mm]"],
        "heated_length_mm": df["length [mm]"],
        "CHF_kW_m2": df["chf_exp [MW/m2]"] * 1000,
        "assumptions": "fluid assumed water (not stated in source file; standard for this literature compilation)",
    })


def load_kaeri_uniform() -> pd.DataFrame:
    df = pd.read_csv(DATA_EXTERNAL / "kaeri_tr1665_uniform_chf.csv")
    return pd.DataFrame({
        "source_row_id": df["TestID"].astype(str),
        "geometry_family": "tube",
        "data_type": "raw_experimental",
        "fluid": df["Fluid"],
        "pressure_kPa": df["Pressure"] / 1000,
        "mass_flux_kg_m2s": df["MassFlux"],
        "quality": df["EquilibriumQuality"],
        "diameter_mm": df["Diameter"] * 1000,
        "heated_length_mm": df["Length"] * 1000,
        "CHF_kW_m2": df["HeatFlux"] / 1000,
        "reference_id": df["Source"],
        "wallpower": df["WallPower"],
        "wallmesh": df["WallMesh"],
    })


def load_kaeri_nonuniform() -> pd.DataFrame:
    df = pd.read_csv(DATA_EXTERNAL / "kaeri_tr1665_nonuniform_chf.csv")
    return pd.DataFrame({
        "source_row_id": df["TestID"].astype(str),
        "geometry_family": "tube",
        "data_type": "raw_experimental",
        "fluid": df["Fluid"],
        "pressure_kPa": df["Pressure"] / 1000,
        "mass_flux_kg_m2s": df["MassFlux"],
        "quality": df["Quality"],
        "diameter_mm": df["Diameter"] * 1000,
        "heated_length_mm": df["Length"] * 1000,
        "CHF_kW_m2": df["HeatFlux"] / 1000,
        "reference_id": df["Source"],
        "chf_location": df["CHFLocation"],
        "shape": df["Shape"],
        "continuous": df["Continuous"],
        "wallpower": df["WallPower"],
        "wallmesh": df["WallMesh"],
    })


def load_pinfin() -> pd.DataFrame:
    # NOTE: "CHF Dataset.csv" is a byte-identical duplicate of this file and is
    # deliberately not loaded separately.
    df = pd.read_csv(DATA_EXTERNAL / "pinfin_chf_water_fc72.csv", encoding="utf-8-sig")
    return pd.DataFrame({
        "source_row_id": df.index.astype(str),
        "geometry_family": "pin_fin_pool_boiling",
        "data_type": "raw_experimental",
        "fluid": df["Fluid"],
        "mass_flux_kg_m2s": 0.0,
        "subcooling_K": df["Subcooling(K)"],
        "fin_shape": df["Fin Shape"],
        "fin_array": df["Fin Array"],
        "fin_width_um": df["Width(um)"],
        "fin_height_um": df["Height(um)"],
        "fin_spacing_um": df["Spacing(um)"],
        "coverage": df["Coverage"],
        "porosity": df["Porosity"],
        "mbl_lateral_um": df["MBL, Lateral (um)"],
        "mbl_total_um": df["MBL, Total (um)"],
        "roughness_factor": df["Roughness Factor"],
        "surface_material": df["Surface Material"],
        "source_citation": df["Source"],
        "CHF_kW_m2": df["CHF(kW/m2)"],
        "assumptions": "pressure not reported in source file; pool boiling, treated as unspecified/atmospheric",
    })


def load_helical_coil() -> pd.DataFrame:
    df = pd.read_csv(DATA_EXTERNAL / "helical_coil_r123_appendixCD.csv")
    return pd.DataFrame({
        "source_row_id": df["SN"].astype(str),
        "geometry_family": "helical_coil",
        "data_type": "raw_experimental",
        "fluid": "R123",
        "pressure_kPa": df["Psys_bar"] * 100,
        "mass_flux_kg_m2s": df["G_kg_m2s"],
        "quality": df["xe"],
        "heated_length_mm": df["Lh_mm"],
        "CHF_kW_m2": df["CHF_kW_m2"],
        "coil_no": df["Coil_no"],
        "appendix": df["appendix"],
        "rho_l_over_rho_g": df["rho_l_over_rho_g"],
        "assumptions": "diameter not present in source CSV (likely in source PDF appendix, not digitized here)",
    })


def load_nureg_sample() -> pd.DataFrame:
    df = pd.read_csv(DATA_EXTERNAL / "nureg_km0011_table4-1_SAMPLE_ONLY.csv")
    for c in ["G_kg_m2s", "CHF_kW_m2"]:
        df[c] = df[c].astype(str).str.replace(",", "", regex=False).astype(float)
    return pd.DataFrame({
        "source_row_id": df["DataID"].astype(str),
        "geometry_family": "tube",
        "data_type": "raw_experimental",
        "fluid": "water",
        "pressure_kPa": df["P_kPa"],
        "mass_flux_kg_m2s": df["G_kg_m2s"],
        "quality": df["Xchf"],
        "diameter_mm": df["D_m"] * 1000,
        "heated_length_mm": df["L_m"] * 1000,
        "inlet_temperature_C": df["Tin_C"],
        "subcooling_kJkg": df["DHin_kJkg"],
        "reference_id": df["Reference"],
        "CHF_kW_m2": df["CHF_kW_m2"],
        "assumptions": "small SAMPLE_ONLY extract (21 rows) -- not the full source dataset, low weight for training",
    })


def load_mentor() -> pd.DataFrame:
    """Extracts the mentor pool-boiling workbook using the same fixed cell
    positions that were manually audited against the workbook's colored
    ("green") cells (see project history / prior plan2 audit). Re-implemented
    here standalone so this pipeline folder has no dependency on old scripts.
    """
    from openpyxl import load_workbook

    path = DATA_RAW / "mentor_master_experiments.xlsx"
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Final Master file"]
    cols = {"Angle": 3, "L_effective_mm": 5, "Width_mm": 6, "Pnet": 20,
            "CHF_MW_m2": 21, "Tsat_Tpool": 26, "Surface_tension": 27,
            "rho_l": 28, "Cp": 41, "Kl": 42, "l_w": 43, "mu_l": 44,
            "alpha": 46, "Ja": 47, "R": 48, "orientation": 2}

    rows = []
    for r, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not values or values[0] is None:
            continue
        row = {k: values[c - 1] if c - 1 < len(values) else None for k, c in cols.items()}
        row["source_row_id"] = str(values[0])
        if row["CHF_MW_m2"] is not None:
            rows.append(row)

    df = pd.DataFrame(rows)
    numeric_cols = [c for c in cols if c != "orientation"]
    for c in numeric_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    return pd.DataFrame({
        "source_row_id": df["source_row_id"],
        "geometry_family": "flat_heater_pool_boiling",
        "data_type": "raw_experimental",
        "fluid": "water",
        "mass_flux_kg_m2s": 0.0,
        "subcooling_K": df["Tsat_Tpool"],
        "heated_length_mm": df["L_effective_mm"],
        "heater_width_mm": df["Width_mm"],
        "orientation": df["orientation"],
        "angle_deg": df["Angle"],
        "tsat_minus_tpool_K": df["Tsat_Tpool"],
        "surface_tension_N_m": df["Surface_tension"],
        "rho_l_kg_m3": df["rho_l"],
        "cp_l_J_kgK": df["Cp"],
        "kl_W_mK": df["Kl"],
        "mu_l_Pa_s": df["mu_l"],
        "alpha_m2_s": df["alpha"],
        "ja": df["Ja"],
        "r_bubble_m": df["R"],
        "CHF_kW_m2": df["CHF_MW_m2"] * 1000,
        "assumptions": "diameter not applicable (flat rectangular heater, not a tube); fluid assumed water",
    })


# NOTE: nureg_km0011_table4-1_SAMPLE_ONLY.csv is intentionally excluded from
# SOURCE_LOADERS. Verified during merge validation: all 21 of its rows are
# exact-value duplicates already present in nrc_groeneveld_24579pt (both
# digitize the same Lowdermilk, 1958 tube-burnout experiment). Including it
# as a second source would let the same physical data point appear twice
# under two different source_dataset labels -- a direct leakage risk for the
# surface-wise / leave-one-source-out splits. load_nureg_sample() is kept
# here for documentation/reference but is not registered below.
SOURCE_LOADERS = {
    "nrc_groeneveld_24579pt": load_nrc_groeneveld,
    "zhao2020": load_zhao2020,
    "kaeri_uniform": load_kaeri_uniform,
    "kaeri_nonuniform": load_kaeri_nonuniform,
    "pinfin_chf_water_fc72": load_pinfin,
    "helical_coil_r123": load_helical_coil,
    "mentor_master": load_mentor,
}
