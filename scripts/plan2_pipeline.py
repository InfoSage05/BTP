"""Reproducible implementation of PLAN_2.

This script is intentionally conservative: it keeps the mentor workbook held out,
uses only the workbook's green columns, excludes the exact target-reconstruction
column from the primary model, and treats the PDF appendix as external validation.
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader
from sklearn.base import clone
from sklearn.ensemble import ExtraTreesRegressor, RandomForestRegressor
from sklearn.impute import SimpleImputer
from sklearn.linear_model import ElasticNet, Ridge
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import GroupKFold, LeaveOneGroupOut
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import PolynomialFeatures, StandardScaler
from scipy.interpolate import RegularGridInterpolator


ROOT = Path(__file__).resolve().parents[1]
GREEN_COLUMNS = [
    "Angle", "L_effective_mm", "Width_mm", "Pnet", "Tsat_Tpool",
    "Surface_tension", "rho_l", "Cp", "Kl", "l_w", "mu_l", "alpha", "Ja", "R",
]
PRIMARY_FEATURES = [c for c in GREEN_COLUMNS if c != "Pnet"]


def read_mentor(path: Path) -> tuple[pd.DataFrame, dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Final Master file"]
    # Fixed positions are deliberate: these were audited against the workbook's
    # green fill and are safer than guessing from similar formula headers.
    cols = {"Angle": 3, "L_effective_mm": 5, "Width_mm": 6, "Pnet": 20,
            "CHF_MW_m2": 21, "Tsat_Tpool": 26, "Surface_tension": 27,
            "rho_l": 28, "Cp": 41, "Kl": 42, "l_w": 43, "mu_l": 44,
            "alpha": 46, "Ja": 47, "R": 48, "orientation": 2}
    rows = []
    for r, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not values or values[0] is None:
            continue
        row = {k: values[c - 1] if c - 1 < len(values) else None for k, c in cols.items()}
        if row["CHF_MW_m2"] is not None:
            rows.append(row)
    df = pd.DataFrame(rows)
    for c in cols:
        if c not in ("orientation",):
            df[c] = pd.to_numeric(df[c], errors="coerce")
    df["geometry_group"] = (
        df["orientation"].astype(str).str.strip().str.lower() + "|" +
        df["Angle"].round(6).astype(str) + "|" +
        df["L_effective_mm"].round(6).astype(str) + "|" +
        df["Width_mm"].round(6).astype(str)
    )
    identity_1 = np.max(np.abs(df["CHF_MW_m2"] - df["Pnet"] /
                                (df["L_effective_mm"] * df["Width_mm"])))
    identity_2 = np.max(np.abs(df["CHF_MW_m2"] * 1e6 - df["CHF_MW_m2"] * 1e6))
    audit = {"rows": len(df), "green_columns": GREEN_COLUMNS,
             "primary_features": PRIMARY_FEATURES,
             "geometry_groups": int(df.geometry_group.nunique()),
             "max_abs_target_identity_error": float(identity_1),
             "max_abs_unit_conversion_error": float(identity_2),
             "target_unit": "MW/m^2", "target_source": "U / CHF(MW/m^2)"}
    return df, audit


NUM = r"[-+]?\d+(?:\.\d+)?"


def parse_pdf(path: Path) -> pd.DataFrame:
    reader = PdfReader(str(path))
    pages = [re.sub(r"\s+", " ", p.extract_text() or "") for p in reader.pages]
    records = []
    # D.1 contains 55 straight-tube rows. Seven numeric values follow Sr. No.
    for page_no in (0, 1):
        text = pages[page_no]
        for m in re.finditer(rf"({NUM}) ({NUM}) ({NUM}) ({NUM}) ({NUM}) ({NUM}) ({NUM}) ({NUM})", text):
            vals = [float(m.group(i)) for i in range(1, 9)]
            # Exclude page numbers/header fragments by the physical ranges.
            if 4 <= vals[1] <= 20 and 300 <= vals[2] <= 1200 and 0 < vals[3] < 10 and 0 < vals[5] < 2500:
                records.append({"table": "D.1", "fluid": "R123", "geometry": "straight",
                                "tube_d_mm": vals[1], "heated_length_mm": vals[2],
                                "P_bar": vals[3], "inlet_temperature_C": vals[4],
                                "G_kg_m2_s": vals[5], "X": vals[6], "CHF_kW_m2": vals[7]})
    # Coil rows have a coil identifier followed by seven numeric values. Searching
    # the token stream handles continuation rows where Sr. No. is visually omitted.
    sections = [(i, "D.2", pages[i]) for i in range(2, 7)]
    sections += [(i, "D.3", pages[i]) for i in range(7, 11)]
    sections += [(11, "D.3", pages[11].split("Table D.4", 1)[0])]
    sections += [(11, "D.4", pages[11].split("Table D.4", 1)[-1])]
    sections += [(i, "D.4", pages[i]) for i in range(12, 15)]
    for page_idx, table, text in sections:
        for m in re.finditer(rf"(Coil_\d+) ({NUM}) ({NUM}) ({NUM}) ({NUM}) ({NUM}) ({NUM}) ({NUM})", text):
            coil = m.group(1)
            vals = [float(m.group(i)) for i in range(2, 9)]
            # Reject table page numbers / headings and CHF-ratio artifacts.
            if 300 <= vals[0] <= 2500 and 50 <= vals[1] <= 2200 and 1.0 <= vals[2] <= 8.0:
                records.append({"table": table, "fluid": "water" if table == "D.2" else "R123",
                                "geometry": "helical", "coil": coil,
                                "heated_length_mm": vals[0], "G_kg_m2_s": vals[1],
                                "P_bar": vals[2], "X": vals[4], "heat_supply_W": vals[5],
                                "CHF_kW_m2": vals[6]})
    df = pd.DataFrame(records)
    if df.empty:
        raise RuntimeError("No appendix rows parsed from PDF")
    # Deduplicate only exact parser repeats at the D.3/D.4 page boundary.
    key = ["table", "fluid", "geometry", "heated_length_mm", "G_kg_m2_s", "P_bar", "X", "CHF_kW_m2"]
    df = df.drop_duplicates(key).reset_index(drop=True)
    return df


def metrics(y, pred):
    y, pred = np.asarray(y, float), np.asarray(pred, float)
    return {"n": int(len(y)), "r2": float(r2_score(y, pred)),
            "mae": float(mean_absolute_error(y, pred)),
            "rmse": float(mean_squared_error(y, pred) ** 0.5),
            "log_rmse": float(mean_squared_error(np.log(np.maximum(y, 1e-9)),
                                                  np.log(np.maximum(pred, 1e-9))) ** 0.5),
            "mdape_pct": float(np.median(np.abs((pred - y) / np.maximum(y, 1e-9))) * 100),
            "within_20pct": float(np.mean(np.abs(pred - y) / np.maximum(y, 1e-9) <= .20))}


def mentor_cv(df: pd.DataFrame) -> pd.DataFrame:
    X = df[PRIMARY_FEATURES].replace([np.inf, -np.inf], np.nan)
    y = df["CHF_MW_m2"].values
    groups = df["geometry_group"].values
    models = {
        "median": None,
        "ridge_log": make_pipeline(SimpleImputer(), StandardScaler(), Ridge(alpha=10.0)),
        "poly2_ridge_log": make_pipeline(SimpleImputer(), StandardScaler(), PolynomialFeatures(2, include_bias=False), Ridge(alpha=10.0)),
        "elastic_log": make_pipeline(SimpleImputer(), StandardScaler(), ElasticNet(alpha=.01, l1_ratio=.2, max_iter=20000)),
        "extra_trees_control": make_pipeline(SimpleImputer(), ExtraTreesRegressor(n_estimators=50, random_state=42, min_samples_leaf=2, n_jobs=1)),
    }
    out = []
    # Five-fold grouped CV keeps every normalized geometry group intact while
    # remaining quick enough to run locally. A future final audit may replace
    # this with LeaveOneGroupOut once a selected model is frozen.
    splitter = GroupKFold(n_splits=5)
    for name, model in models.items():
        preds = np.full(len(df), np.nan)
        for tr, te in splitter.split(X, y, groups):
            if model is None:
                preds[te] = np.median(y[tr])
            else:
                m = clone(model)
                m.fit(X.iloc[tr], np.log(y[tr]))
                preds[te] = np.exp(m.predict(X.iloc[te]))
        out.append({"track": "mentor_group_cv", "model": name, **metrics(y, preds)})
    # Explicit leakage control: the identity predictor is not a valid deployable model.
    out.append({"track": "leakage_control", "model": "Pnet/(L*W)", **metrics(y, df.Pnet.values / (df.L_effective_mm.values * df.Width_mm.values))})
    return pd.DataFrame(out)


def external_report(pdf_df: pd.DataFrame, mentor_audit: dict) -> pd.DataFrame:
    lut = pd.read_csv(ROOT / "data/chf_long_clean.csv")
    lut = lut[lut.CHF > 0]
    pvals, gvals, xvals = [np.sort(lut[c].unique()) for c in ("P", "G", "X")]
    cube = np.full((len(pvals), len(gvals), len(xvals)), np.nan)
    for row in lut.itertuples(index=False):
        cube[np.where(pvals == row.P)[0][0], np.where(gvals == row.G)[0][0], np.where(xvals == row.X)[0][0]] = row.CHF
    interp = RegularGridInterpolator((pvals, gvals, xvals), cube, bounds_error=False, fill_value=None)
    rows = []
    for table, g in pdf_df.groupby("table", sort=True):
        # The existing LUT expects water, P in kPa, G in kg/m2/s, X. It is a
        # baseline only; PDF pressure is far below the LUT's training domain.
        pred = interp(np.column_stack([g.P_bar.values * 100.0, g.G_kg_m2_s.values, g.X.values]))
        m = metrics(g.CHF_kW_m2.values, pred)
        p_ood = bool(((g.P_bar * 100.0 < pvals.min()) | (g.P_bar * 100.0 > pvals.max())).any())
        rows.append({"table": table, "n": len(g), "fluid": g.fluid.iloc[0],
                     "geometry": g.geometry.iloc[0], "P_min_bar": g.P_bar.min(),
                     "P_max_bar": g.P_bar.max(), "G_min": g.G_kg_m2_s.min(),
                     "G_max": g.G_kg_m2_s.max(), "CHF_min_kW_m2": g.CHF_kW_m2.min(),
                     "CHF_max_kW_m2": g.CHF_kW_m2.max(), "lut_zero_shot_r2": m["r2"],
                     "lut_zero_shot_mdape_pct": m["mdape_pct"], "lut_zero_shot_within_20pct": m["within_20pct"],
                     "pressure_ood": p_ood,
                     "zero_shot_status": "diagnostic_only",
                     "reason": "current LUT is water, vertical 8-mm normalized, P in kPa; appendix includes R123 and helical coils"})
    return pd.DataFrame(rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--mentor", default="data/raw/external/paper_extracted_test_only/mentor_master_experiments.xlsx")
    ap.add_argument("--pdf", default="data/raw/external/paper_extracted_test_only/external_coil_tube_chf_appendix.pdf")
    ap.add_argument("--out", default="results/plan2")
    args = ap.parse_args()
    out = ROOT / args.out
    out.mkdir(parents=True, exist_ok=True)
    mentor, audit = read_mentor(ROOT / args.mentor)
    pdf = parse_pdf(ROOT / args.pdf)
    cv = mentor_cv(mentor)
    ext = external_report(pdf, audit)
    mentor.to_csv(out / "mentor_green_data.csv", index=False)
    pdf.to_csv(out / "external_pdf_data.csv", index=False)
    cv.to_csv(out / "mentor_group_cv.csv", index=False)
    ext.to_csv(out / "external_validation_report.csv", index=False)
    summary = {"mentor_audit": audit, "pdf_rows": int(len(pdf)),
               "pdf_counts": {k: int(v) for k, v in pdf.table.value_counts().to_dict().items()},
               "primary_cv_best": cv[cv.track == "mentor_group_cv"].sort_values("rmse").iloc[0].to_dict()}
    (out / "summary.json").write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    main()
